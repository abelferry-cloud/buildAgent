"""Excel generation skill using openpyxl."""

import os
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    Workbook = None

from agent.tools.base import Tool, ToolResult


class ExcelGenerateTool(Tool):
    """Generate an Excel spreadsheet."""

    name = "excel_generate"
    description = "Generate an Excel spreadsheet with data."

    def __init__(self, output_dir: str = "output"):
        super().__init__()
        self.output_dir = os.path.join(os.getcwd(), output_dir)
        os.makedirs(self.output_dir, exist_ok=True)

    def execute(
        self,
        headers: list[str],
        rows: list[list[str]],
        sheet_name: str = "Sheet1",
        output_filename: str = "spreadsheet.xlsx",
    ) -> ToolResult:
        """Generate an XLSX file."""
        if Workbook is None:
            return ToolResult(
                output="",
                error="openpyxl not installed. Run: pip install openpyxl"
            )

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            output_path = os.path.join(self.output_dir, output_filename)
            wb.save(output_path)
            return ToolResult(output=f"Excel 文件已生成: {output_path}")
        except Exception as e:
            return ToolResult(output="", error=str(e))
